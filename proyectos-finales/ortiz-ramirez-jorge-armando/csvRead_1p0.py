#-------------------------------------------------------------------------------
#
#
# Curso: Python Básico Intel
# Autor: Jorge Armando Ortiz Ramirez
# Correo: jaor87@gmail.com
#
# Proyecto Final.
#
# Problemas que se resuelven:
#   8. Crear una gráfica de Serie
#   9. Cargar un dataframe de datos desde un CSV
#   xx. Guardar imágenes en formato .png en un archivo de Excel.
#
#
# Justificación:
#   Actualmente los archivos en formato CSV son muy utilizados para guardar
#   y manipular una gran cantidad de datos, por ejemplo los osciloscopios de
#   keysight pueden guardar informacion en dicho formato. Poder manipular y
#   gráficar dicha información de manera automatizada y además tener la opción
#   de generar un reporte resulta muy útil tanto en el ámbito laboral como
#   cientifico.
#
#   En el presente código se pretende tomar información de un CSV previamente
#   generado, extraer la información, graficarla y guardar la gráfica en formato
#   .png para tener la posibilidad de agregarla en un reporte en un archivo de
#   Excel.
#
#
# Pequeña descripción:
#   1. Mediante la librería pathlib se accede a un directorio especificado y se
#      busca un patron en los nombres de los archivos que se encuentran en dicho
#      directorio, ésto último haciendo uso de la librería re.
#
#   2. Se extrae la informacion de dicho CSV y se carga en un dataframe con
#      ayuda de la librería pandas.
#
#   3. Se grafica con pyplot la información del CSV o los CSVs que coincidieron
#      con el patrón de búsqueda y se muestra.
#
#   4. Si así se desea, se agrega la imágen de la gráfica en formato .png a un
#      archivo de Excel.
#
#
# Librerias de python necesarias:
#   pandas
#   matplotlib
#   time
#   pathlib
#   re
#   openpyxl
#
#-------------------------------------------------------------------------------

import pandas as pd
import time
import matplotlib.pyplot as plt
import pathlib
import re
import openpyxl

class csvPlot:
    # Atributos
    path=r"C:\CursoPythonBasico\Proyecto\folderCsvs"
    pattern1 = "PS0"   # patron a reconocer en el nombre del archivo
    pattern2 = "33"  # patron a reconocer en el nombre del archivo

    # Métodos
    def csvList(self):
        """
        Se extraen los nombres de todos los archivos CSVs existentes en el
        atributo directorio 'path' y que cumplen con ciertas
        caracteristicas en su nombre -- pattern1 y pattern2.

        Regresa -- Una lista con dichos nombres
        """
        pattern = r".*{}.*{}C.*.csv$".format(self.pattern1, self.pattern2)  # loquesea + <pattern1> + loquesea + <pattern2> + loquesea + termina con .csv

        FileNamesList=[]
        dirPath = self.path
        directory = pathlib.Path(dirPath)
        #print("Filtro: ")
        for fichero in directory.iterdir():     # fichero es un objeto
            nameStr=fichero.name                 # fichero.name es un dato tipo str
            #print("No filtro: ",nameStr)
            fileName=re.match(pattern,nameStr)   # si no hay coincidencia con el patron, re.match regresa None
            if fileName is not None:
                FileNamesList.append(fileName.group(0))
        return FileNamesList

    def csvToDataframe(self, csvName):
        """
        Toma un CSV de formato especifico y lo comvierte en Data frame.
        Usa el atributo path para encontrar el nombre del archivo especificado.

        Argumentos:
            csvName -- str, nombre del CSV.
        """
        pathCsv = "{}\\{}".format(self.path, csvName)
        print(pathCsv)
        df = pd.read_csv(pathCsv, skiprows=37) # saltar las primeras 37 lineas del CSV
        return df

    def plot(self, plotName="Fig", makeReport = False):
        """
        Este metodo grafica los datos de todos los CSVs que se especifican con
        el path y los patrones dados.
        Guarda la gráfica en una imágen con formato .png bajo el nombre
        especificado.
        Es opcional si se quiere generar un reporte con la gráfica generada.

        Parametros:
            plotName --> tipo str, nombre que se le va a dar a la imágen con la gráfica.
            makeReport --> tipo bool, True si se quiere generar el reporte en Excel
        """
        #fig=plt.figure()
        d = time.asctime()
        date= d.split(' ')[1]+d.split(' ')[2]+'_'+d.split(' ')[3].replace(':','-') # example 'Sep_22_18-27-29'
        nameFig = "{}_{}C_{}_{}.png".format(self.pattern1, self.pattern2, plotName, date) # nombre de la figura
        fig, ax = plt.subplots(1, 1, constrained_layout=True) # (2, 1, 1)
        fig.suptitle(f'{nameFig}', fontsize = 14)

        for csv in self.csvList():
            df = self.csvToDataframe(csv)
            file=df                        # el df puede ser r"C:\CursoPythonBasico\Proyecto\folderCsvs\Base_PS0_50C_1p0V_0x3cpspwr_0x0df_csx.csv"  #
            frecuencias=file['!Frequency (Hz)']
            amplitud = file[' Amplitude']

            lfrequencias =frecuencias.to_list()
            lamplitud = amplitud.to_list()
            ax.plot(lfrequencias, lamplitud, lw=1.25, label= csv)
            ax.set_title("Data to show",fontname='Comic Sans MS', fontsize=18, color='purple')
            ax.set_label(csv)
            ax.set_xlabel('data x Units')
            ax.set_ylabel('data y Units')
            ax.grid(which='both', linestyle='-')
        plt.show()#pylab.show()
        time.sleep(1)

        print("Nombre de Figura: ", nameFig)
        saveAs = "{}\\{}_{}C_{}_{}.png".format(self.path, self.pattern1, self.pattern2, plotName, date) # directorio de la figura con nombre
        fig.savefig(saveAs, dpi=150)
        time.sleep(0.5)

        if makeReport:
            i=5
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.title = "{}C".format(nameFig.split('C')[0])
            print(f'Hoja activa: {wb.active.title}')
            sheet["A1"] = "Proyecto Final, Reporte con grágicas en una misma imágen."
            sheet["A2"] = "Python básico, Jorge Ortiz"

            sheet.cell(row=i-1, column=1, value="CSV graficados:")
            for name in self.csvList():
                sheet.cell(row=i, column=2, value=name)
                i=i+1
            sheet["A12"] = "Imágen:"
            sheet["B12"] = nameFig
            img = openpyxl.drawing.image.Image(f'{saveAs}')
            img.anchor = 'A13'
            sheet.add_image(img)
            nameExcel = 'FinalReport_{}.xlsx'.format(date)
            wb.save('{}\\{}'.format(self.path, nameExcel))
            print('Resultados guardados en: {}'.format(nameExcel))

        return saveAs # regresa el directorio completo


Graficar = csvPlot()
Graficar.path        = r"C:\CursoPythonBasico\Proyecto\folderCsvs"
Graficar.pattern1   = "PS0"
Graficar.pattern2 = "33"
Graficar.plot(plotName="Plots", makeReport=True)




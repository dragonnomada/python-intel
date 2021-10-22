#-------------------------------------------------------------------------------
# Name:        Efficiency report
# Purpose:     Generate PDF report from an excel file
#
# Motivation:  In Validation, it is very important to collect all kinds of data in
#              order to verify the correct behavior of certain cicuits. This data
#              is often overwhelming since it can be hundreds of measurements.
#              designers and co-workers find very difficult to understand this data.
#              This is a project that wants to facilitate the understanding of the
#              data through a chart.
#
# Description: This project takes data from a excel file using pandas and openpyxl
#              and stores in 2 lists the values corresponding to each point of
#              measurement. Using matplotlib I created a chart that represents
#              the efficiency. I did a report with module fpdf using txt
#              archives and the plot from matplotlib.
#
# Author:      Karla Gabriela Lopez Araiza
#
# Created:     18/10/2021
# Copyright:   (c) Karla Lopez 2021
#
#-------------------------------------------------------------------------------
import pandas as pd
import openpyxl
import matplotlib.pyplot as plt
from fpdf import FPDF

def Efficiency():
 Current=[]
 Efficiency=[]
 iterator=0
 df_Efficiency_sheet = pd.read_excel("C:\Documents\Efficiency.xlsx", sheet_name="Efficiency", engine="openpyxl", usecols="A:E")

 print(df_Efficiency_sheet)

 for take in df_Efficiency_sheet["Measurement"]:
    data = df_Efficiency_sheet.query(f"Measurement == '{take}'")

    Current.append(round(data["OutCurrent"][iterator],4))
    Efficiency.append(round(data["EffOut"][iterator]*100,4))

    iterator= iterator+1

 font1 = {'family':'serif','color':'black','size':20}

 plt.plot(Current, Efficiency, "ro-")
 plt.xlabel("Current (A)")
 plt.ylabel("Efficiency (%)")
 plt.title("Efficiency Measurements", fontdict = font1)

 plt.savefig("C:\Documents\Efficiency.png")

 pdf = FPDF("P", "pt", "A4")

 pdf.add_page()

 pdf.set_font("Times", "I", 15)

 pdf.cell(0, 4, "Python Basico")

 pdf.ln(20)

 pdf.cell(0, 4, "Karla Lopez")

 pdf.ln(30)

 pdf.set_font("Times", "BI", 25)

 pdf.set_text_color(220, 50, 50)

 pdf.cell(0, 4, "Reporte de Eficiencia")

 pdf.ln(40)

 pdf.image("C:\Documents\Efficiency.png", x=100, w=400)

 pdf.ln(20)

 f=open("C:\Documents\Reporte.txt","r")

 contenido = f.read()

 f.close()

 pdf.set_font("Times", "", 11)

 pdf.set_text_color(0, 0, 0)

 pdf.multi_cell(0, 4, contenido)

 pdf.output("C:\Documents\Reporte.pdf")



if __name__ == '__main__':
    Efficiency()
